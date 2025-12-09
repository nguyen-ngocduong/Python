import numpy as np
import openpyxl
import os

# Ensure data directory exists
if not os.path.exists('data'):
    os.makedirs('data')

# =============================================================================
# THAO TÁC DỮ LIỆU EXCEL VỚI NUMPY
# =============================================================================

print("=" * 70)
print("THAO TÁC DỮ LIỆU EXCEL VỚI NUMPY")
print("=" * 70)

# 1. TẠO FILE EXCEL MẪU
print("\n1. TẠO FILE EXCEL MẪU")
print("-" * 40)

# Tạo dữ liệu nhân viên mẫu
nhan_vien_data = [
    ['ID', 'Ho_Ten', 'Tuoi', 'Luong', 'Phong_Ban', 'Kinh_Nghiem'],
    [1, 'Nguyen Van A', 28, 15000000, 'IT', 3],
    [2, 'Tran Thi B', 32, 18000000, 'HR', 5],
    [3, 'Le Van C', 25, 12000000, 'IT', 2],
    [4, 'Pham Thi D', 35, 22000000, 'Finance', 8],
    [5, 'Hoang Van E', 29, 16000000, 'IT', 4],
    [6, 'Vu Thi F', 31, 19000000, 'HR', 6],
    [7, 'Dao Van G', 27, 14000000, 'Marketing', 3],
    [8, 'Bui Thi H', 33, 20000000, 'Finance', 7],
    [9, 'Ngo Van I', 26, 13000000, 'IT', 2],
    [10, 'Ly Thi K', 30, 17000000, 'Marketing', 5]
]
ban_hang_data = np.array([
    [1, 150, 2250000, 8.5],
    [2, 200, 3000000, 9.2],
    [3, 175, 2625000, 7.8],
    [4, 220, 3300000, 9.5],
    [5, 180, 2700000, 8.1],
    [6, 195, 2925000, 8.9],
    [7, 165, 2475000, 7.5],
    [8, 210, 3150000, 9.1],
    [9, 185, 2775000, 8.3],
    [10, 225, 3375000, 9.7]
])

# Ghi file Excel cho nhan_vien
wb_nhan_vien = openpyxl.Workbook()
ws_nhan_vien = wb_nhan_vien.active
ws_nhan_vien.title = "NhanVien"
for row in nhan_vien_data:
    ws_nhan_vien.append(row)
wb_nhan_vien.save('data/nhan_vien.xlsx')

# Ghi file Excel cho ban_hang
wb_ban_hang = openpyxl.Workbook()
ws_ban_hang = wb_ban_hang.active
ws_ban_hang.title = "BanHang"
header = ['Thang', 'So_Luong_Ban', 'Doanh_Thu', 'Danh_Gia']
ws_ban_hang.append(header)
for row in ban_hang_data:
    ws_ban_hang.append(row.tolist())
wb_ban_hang.save('data/ban_hang.xlsx')

print("✓ Đã tạo file 'nhan_vien.xlsx' và 'ban_hang.xlsx'")

# 2. ĐỌC FILE EXCEL VỚI NUMPY
print("\n2. ĐỌC FILE EXCEL VỚI NUMPY")
print("-" * 40)

wb_ban_hang = openpyxl.load_workbook('data/ban_hang.xlsx')
ws_ban_hang = wb_ban_hang['BanHang']
ban_hang = np.array([row for row in ws_ban_hang.iter_rows(min_row=2, values_only=True)], dtype=float)
print("Dữ liệu bán hàng:")
print(ban_hang)
print(f"Kích thước: {ban_hang.shape}")

wb_nhan_vien = openpyxl.load_workbook('data/nhan_vien.xlsx')
ws_nhan_vien = wb_nhan_vien['NhanVien']
nv_so_lieu = np.array([
    (row[0], row[2], row[3], row[5])
    for row in ws_nhan_vien.iter_rows(min_row=2, values_only=True)
], dtype=float)
print(f"\nDữ liệu số của nhân viên:")
print("Cột: [ID, Tuoi, Luong, Kinh_Nghiem]")
print(nv_so_lieu)

try:
    nhan_vien = np.array([
        tuple(row)
        for row in ws_nhan_vien.iter_rows(min_row=2, values_only=True)
    ], dtype=object)
    print(f"\nDữ liệu nhân viên (5 dòng đầu):")
    for record in nhan_vien[:5]:
        print(f"  {record}")
except:
    print("Lưu ý: Đọc file có text phức tạp, đã xử lý riêng từng cột số")

# 3. PHÂN TÍCH DỮ LIỆU BÁN HÀNG
print("\n3. PHÂN TÍCH DỮ LIỆU BÁN HÀNG")
print("-" * 40)

# Trích xuất các cột
thang = ban_hang[:, 0]
so_luong = ban_hang[:, 1]
doanh_thu = ban_hang[:, 2]
danh_gia = ban_hang[:, 3]

print(f"Tháng: {thang}")
print(f"Số lượng bán: {so_luong}")
print(f"Doanh thu: {doanh_thu}")
print(f"Đánh giá: {danh_gia}")

# Thống kê cơ bản
print("\nTHỐNG KÊ BÁN HÀNG:")
print(f"📊 Tổng số lượng bán: {np.sum(so_luong):,.0f}")
print(f"💰 Tổng doanh thu: {np.sum(doanh_thu):,.0f} VNĐ")
print(f"📈 Doanh thu trung bình: {np.mean(doanh_thu):,.0f} VNĐ")
print(f"🔝 Doanh thu cao nhất: {np.max(doanh_thu):,.0f} VNĐ")
print(f"🔻 Doanh thu thấp nhất: {np.min(doanh_thu):,.0f} VNĐ")
print(f"⭐ Đánh giá trung bình: {np.mean(danh_gia):.2f}/10")

# Tìm tháng có doanh thu cao nhất
thang_max = thang[np.argmax(doanh_thu)]
print(f"🏆 Tháng có doanh thu cao nhất: Tháng {thang_max:.0f}")

# 4. PHÂN TÍCH DỮ LIỆU NHÂN VIÊN
print("\n4. PHÂN TÍCH DỮ LIỆU NHÂN VIÊN")
print("-" * 40)

# Trích xuất các cột số
nv_id = nv_so_lieu[:, 0]
nv_tuoi = nv_so_lieu[:, 1]
nv_luong = nv_so_lieu[:, 2]
nv_kinhnghiem = nv_so_lieu[:, 3]

print("THỐNG KÊ NHÂN VIÊN:")
print(f"👥 Tổng số nhân viên: {len(nv_tuoi)}")
print(f"🎂 Tuổi trung bình: {np.mean(nv_tuoi):.1f} tuổi")
print(f"💼 Lương trung bình: {np.mean(nv_luong):,.0f} VNĐ")
print(f"🎯 Kinh nghiệm trung bình: {np.mean(nv_kinhnghiem):.1f} năm")
print(f"💰 Lương cao nhất: {np.max(nv_luong):,.0f} VNĐ")
print(f"💸 Lương thấp nhất: {np.min(nv_luong):,.0f} VNĐ")

# 5. CÁC THAO TÁC NÂNG CAO
print("\n5. CÁC THAO TÁC NÂNG CAO")
print("-" * 40)

# Lọc dữ liệu - Nhân viên có lương > 15 triệu
nv_luong_cao = nv_so_lieu[nv_luong > 15000000]
print(f"Số nhân viên có lương > 15 triệu: {len(nv_luong_cao)}")
print(f"ID nhân viên lương cao: {nv_luong_cao[:, 0]}")

# Sắp xếp theo lương
nv_sorted = nv_so_lieu[np.argsort(nv_luong)]
print(f"\nTop 3 lương thấp nhất:")
for i in range(3):
    print(f"  ID {nv_sorted[i, 0]:.0f}: {nv_sorted[i, 2]:,.0f} VNĐ")

print(f"\nTop 3 lương cao nhất:")
for i in range(-3, 0):
    print(f"  ID {nv_sorted[i, 0]:.0f}: {nv_sorted[i, 2]:,.0f} VNĐ")

# 6. TÍNH TOÁN THỐNG KÊ NÂNG CAO
print("\n6. TÍNH TOÁN THỐNG KÊ NÂNG CAO")
print("-" * 40)

# Tương quan giữa tuổi và lương
correlation = np.corrcoef(nv_tuoi, nv_luong)[0, 1]
print(f"Tương quan tuổi-lương: {correlation:.3f}")

# Tương quan giữa kinh nghiệm và lương
correlation_exp = np.corrcoef(nv_kinhnghiem, nv_luong)[0, 1]
print(f"Tương quan kinh nghiệm-lương: {correlation_exp:.3f}")

# Percentiles
p25 = np.percentile(nv_luong, 25)
p50 = np.percentile(nv_luong, 50)  # median
p75 = np.percentile(nv_luong, 75)

print(f"\nPhân vị lương:")
print(f"  P25 (25%): {p25:,.0f} VNĐ")
print(f"  P50 (50%): {p50:,.0f} VNĐ")
print(f"  P75 (75%): {p75:,.0f} VNĐ")

# 7. GHI DỮ LIỆU ĐÃ XỬ LÝ RA FILE MỚI
print("\n7. GHI DỮ LIỆU ĐÃ XỬ LÝ")
print("-" * 40)

# Tạo báo cáo thống kê
thong_ke_ban_hang = [
    ['Thong_Ke', 'Gia_Tri'],
    ['Tong_So_Luong', np.sum(so_luong)],
    ['Tong_Doanh_Thu', np.sum(doanh_thu)],
    ['TB_Doanh_Thu', np.mean(doanh_thu)],
    ['Max_Doanh_Thu', np.max(doanh_thu)],
    ['Min_Doanh_Thu', np.min(doanh_thu)],
    ['TB_Danh_Gia', np.mean(danh_gia)]
]

# Ghi ra file Excel
wb_bao_cao = openpyxl.Workbook()
ws_bao_cao = wb_bao_cao.active
ws_bao_cao.title = "BaoCaoBanHang"
for row in thong_ke_ban_hang:
    ws_bao_cao.append(row)
wb_bao_cao.save('data/bao_cao_ban_hang.xlsx')

# Ghi dữ liệu nhân viên lương cao
wb_luong_cao = openpyxl.Workbook()
ws_luong_cao = wb_luong_cao.active
ws_luong_cao.title = "NhanVienLuongCao"
header_luong_cao = ['ID', 'Tuoi', 'Luong', 'Kinh_Nghiem']
ws_luong_cao.append(header_luong_cao)
for row in nv_luong_cao:
    ws_luong_cao.append(row.tolist())
wb_luong_cao.save('data/nhan_vien_luong_cao.xlsx')

print("✓ Đã tạo file 'bao_cao_ban_hang.xlsx'")
print("✓ Đã tạo file 'nhan_vien_luong_cao.xlsx'")

# 8. XỬ LÝ DỮ LIỆU THIẾU (MISSING DATA)
print("\n8. XỬ LÝ DỮ LIỆU THIẾU")
print("-" * 40)

# Tạo dữ liệu có giá trị thiếu
du_lieu_thieu = np.array([
    [1, 25, 15000000, 3],
    [2, np.nan, 18000000, 5],
    [3, 30, np.nan, 4],
    [4, 35, 22000000, np.nan],
    [5, 28, 16000000, 6]
])

print("Dữ liệu gốc có giá trị thiếu:")
print(du_lieu_thieu)

# Kiểm tra giá trị thiếu
co_nan = np.isnan(du_lieu_thieu)
print(f"\nVị trí có giá trị thiếu:\n{co_nan}")
print(f"Tổng số giá trị thiếu: {np.sum(co_nan)}")

# Thay thế giá trị thiếu bằng giá trị trung bình
du_lieu_clean = du_lieu_thieu.copy()
for i in range(du_lieu_thieu.shape[1]):
    col = du_lieu_thieu[:, i]
    if np.any(np.isnan(col)):
        mean_val = np.nanmean(col)  # Trung bình bỏ qua NaN
        du_lieu_clean[:, i] = np.where(np.isnan(col), mean_val, col)

print(f"\nDữ liệu sau khi thay thế NaN bằng trung bình:")
print(du_lieu_clean)

# 9. KẾT HỢP NHIỀU FILE EXCEL
print("\n9. KẾT HỢP NHIỀU FILE EXCEL")
print("-" * 40)

# Tạo thêm file Excel khác
du_lieu_them = np.array([
    [11, 24, 11000000, 1],
    [12, 36, 25000000, 10],
    [13, 29, 17000000, 5]
])

wb_them = openpyxl.Workbook()
ws_them = wb_them.active
ws_them.title = "NhanVienThem"
ws_them.append(['ID', 'Tuoi', 'Luong', 'Kinh_Nghiem'])
for row in du_lieu_them:
    ws_them.append(row.tolist())
wb_them.save('data/nhan_vien_them.xlsx')

# Đọc và kết hợp
wb_them = openpyxl.load_workbook('data/nhan_vien_them.xlsx')
ws_them = wb_them['NhanVienThem']
nv_them = np.array([
    tuple(row)
    for row in ws_them.iter_rows(min_row=2, values_only=True)
], dtype=float)
nv_ket_hop = np.vstack([nv_so_lieu, nv_them])

print(f"Dữ liệu gốc: {nv_so_lieu.shape}")
print(f"Dữ liệu thêm: {nv_them.shape}")
print(f"Dữ liệu kết hợp: {nv_ket_hop.shape}")
print(f"\nDữ liệu kết hợp:")
print(nv_ket_hop)

print("\n" + "=" * 70)
print("HOÀN THÀNH! ĐÃ THỰC HIỆN CÁC THAO TÁC EXCEL VỚI NUMPY")
print("Files đã tạo: nhan_vien.xlsx, ban_hang.xlsx, bao_cao_ban_hang.xlsx")
print("nhan_vien_luong_cao.xlsx, nhan_vien_them.xlsx")
print("=" * 70)